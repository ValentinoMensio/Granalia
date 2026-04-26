from __future__ import annotations

import argparse
import os
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_DIR = BASE_DIR / "2026"
DEFAULT_POSTGRES_URL = "postgresql+psycopg://granalia:granalia@127.0.0.1:5432/granalia"
ALLOWED_OFFERINGS = {"16x300 gr", "12x350 gr", "12x400 gr", "10x500 gr", "10x1 kg", "x 4 kg", "x 5 kg", "x 25 kg", "x 30 kg"}


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_cell_text(value: Any) -> str:
    text = str(value or "").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def parse_percent(text: str) -> float | None:
    match = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", text or "")
    if not match:
        return None
    return float(match.group(1).replace(",", ".")) / 100.0


def parse_formula_multiplier(formula: Any) -> float | None:
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    matches = re.findall(r"\*(0?\.\d+|\d+\.\d+)", formula)
    if not matches:
        return None
    try:
        return float(matches[-1])
    except ValueError:
        return None


def discount_key_for_label(label: str) -> str:
    text = normalize_text(label)
    if "16x300" in text:
        return "Pack 300/350/400 gr"
    if "12x300" in text:
        return "Pack 300/350/400 gr"
    if "12x350" in text or "12x400" in text:
        return "Pack 300/350/400 gr"
    if "10x500" in text or "12x500" in text:
        return "Pack 500 gr"
    if "10x1 kg" in text or "10x1000" in text or "10x 1 kg" in text:
        return "Pack 1 kg"
    if "x 4 kg" in text or "x4 kg" in text:
        return "Bolsa 4 kg"
    if "x 5 kg" in text or "x5 kg" in text:
        return "Bolsa 5 kg"
    if "x 25 kg" in text or "x25 kg" in text:
        return "Bolsa 25 kg"
    if "x 30 kg" in text or "x30 kg" in text:
        return "Bolsa 30 kg"
    return clean_cell_text(label) or "Otros"


@dataclass(slots=True)
class ParsedItem:
    label: str
    quantity: int
    unit_price: int
    gross: int
    discount: int
    total: int
    line_discount_rate: float | None = None


@dataclass(slots=True)
class ParsedInvoice:
    path: Path
    client_name: str
    order_date: date
    secondary_line: str = ""
    transport: str = ""
    notes: list[str] = field(default_factory=list)
    footer_discounts: list[dict[str, float | str]] = field(default_factory=list)
    line_discounts_by_format: dict[str, float] = field(default_factory=dict)
    items: list[ParsedItem] = field(default_factory=list)

    @property
    def legacy_key(self) -> str:
        try:
            relative = self.path.relative_to(BASE_DIR)
        except ValueError:
            relative = self.path
        return f"2026:{relative.as_posix()}"


def as_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, str) and value.startswith("="):
        return 0
    return int(round(float(value)))


def as_rate(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        rate = float(value)
        if rate > 1:
            rate /= 100
        return rate if rate > 0 else None
    text = str(value)
    parsed = parse_percent(text)
    if parsed:
        return parsed
    return parse_formula_multiplier(text)


def formula_total(formula: Any, quantity: int, unit_price: int, discount: int, rate: float | None) -> int | None:
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    gross = quantity * unit_price
    normalized = formula.replace(" ", "").lower()
    multiplier = parse_formula_multiplier(formula)
    if multiplier is not None:
        return round(gross * multiplier)
    if "-d" in normalized:
        return gross - discount
    if rate is not None and "*" in normalized:
        return round(gross * rate)
    if "*" in normalized:
        return gross
    return None


def row_text(ws, row: int) -> str:
    values = [clean_cell_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
    return " ".join(value for value in values if value)


def find_row(ws, needle: str) -> int | None:
    normalized_needle = normalize_text(needle)
    for row in range(1, ws.max_row + 1):
        if normalized_needle in normalize_text(row_text(ws, row)):
            return row
    return None


def parse_client_name(value: Any, fallback: Path) -> str:
    text = clean_cell_text(value)
    match = re.search(r"cliente\s*:\s*(.+)", text, flags=re.IGNORECASE)
    if match:
        text = match.group(1)
    text = clean_cell_text(text)
    if text:
        return text
    return re.sub(r"\d{1,2}[- ]\d{1,2}.*$", "", fallback.stem).strip() or fallback.stem


MONTHS_BY_FOLDER = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
}


def parse_date_from_filename(path: Path) -> tuple[date, bool] | None:
    match = re.search(r"(?<!\d)(\d{1,2})[- ](\d{1,2})(?:[- ](20\d{2}))?", path.stem)
    if match:
        day, month, year = match.groups()
        has_explicit_year = year is not None
        parsed_year = int(year) if year else 2026
        try:
            return date(parsed_year, int(month), int(day)), has_explicit_year
        except ValueError:
            return None
    return None


def parse_cell_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_cell_text(value)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_order_date(value: Any, fallback: Path) -> date:
    filename_match = parse_date_from_filename(fallback)
    cell_date = parse_cell_date(value)
    if filename_match:
        filename_date, has_explicit_year = filename_match
        folder_month = MONTHS_BY_FOLDER.get(normalize_text(fallback.parent.name))
        filename_matches_folder = folder_month is None or filename_date.month == folder_month
        if filename_matches_folder and (not has_explicit_year or filename_date.year == 2026):
            return filename_date
        if cell_date:
            return cell_date
        return filename_date
    if cell_date:
        return cell_date
    raise ValueError(f"No se pudo detectar la fecha en {fallback}")


def extract_transport(text: str) -> str:
    cleaned = clean_cell_text(text)
    if not re.search(r"\b(transporte|tpte|transp)\b", cleaned, flags=re.IGNORECASE):
        return ""
    quoted = re.search(r"[\"'“”](.+?)[\"'“”]", cleaned)
    if quoted:
        return clean_cell_text(quoted.group(1))
    cleaned = re.sub(r"\b(transporte|tpte|transp)\b\.?", "", cleaned, flags=re.IGNORECASE)
    return clean_cell_text(cleaned)


def parse_footer(ws, start_row: int) -> tuple[str, list[str], list[dict[str, float | str]]]:
    transport = ""
    notes: list[str] = []
    discounts: list[dict[str, float | str]] = []
    for row in range(start_row, ws.max_row + 1):
        text = row_text(ws, row)
        if not text:
            continue
        normalized = normalize_text(text)
        if "total" in normalized and not re.search(r"dto|desc", normalized):
            continue
        rate = parse_percent(text)
        if rate and re.search(r"dto|desc", normalized):
            discounts.append({"label": f"Dto {round(rate * 100, 2):g}%", "rate": rate})
            continue
        parsed_transport = extract_transport(text)
        if parsed_transport:
            transport = parsed_transport
            continue
        if not re.fullmatch(r"[p\s]+", normalized):
            notes.append(text)
    return transport, notes, discounts


def parse_invoice(path: Path) -> ParsedInvoice:
    workbook = load_workbook(path, data_only=False)
    ws = workbook.active
    products_row = find_row(ws, "PRODUCTOS")
    if not products_row:
        raise ValueError("no se encontro la fila PRODUCTOS")

    client_name = parse_client_name(ws["A7"].value, path)
    order_date = parse_order_date(ws["C6"].value, path)
    secondary_line = ""
    if products_row > 8:
        maybe_secondary = clean_cell_text(ws["A8"].value)
        if maybe_secondary and "producto" not in normalize_text(maybe_secondary):
            secondary_line = maybe_secondary

    header_text = row_text(ws, products_row + 1) + " " + row_text(ws, products_row + 2)
    header_rate = parse_percent(header_text)
    has_line_discount_header = bool(header_rate and re.search(r"dto|desc", normalize_text(header_text)))

    items: list[ParsedItem] = []
    line_discounts: dict[str, float] = {}
    footer_start = ws.max_row + 1
    for row in range(products_row + 3, ws.max_row + 1):
        first = clean_cell_text(ws.cell(row, 1).value)
        normalized_first = normalize_text(first)
        current_row_text = row_text(ws, row)
        normalized_row = normalize_text(current_row_text)
        if first and "total" in normalized_first:
            footer_start = row + 1
            break
        if not first and re.search(r"\b(total|sub-total|subtotal|dto|desc)\b", normalized_row):
            footer_start = row
            break
        quantity = as_int(ws.cell(row, 2).value)
        if not first or quantity <= 0:
            continue
        unit_price = as_int(ws.cell(row, 3).value)

        row_rate = as_rate(ws.cell(row, 4).value)
        if row_rate is None and has_line_discount_header:
            row_rate = header_rate

        gross = quantity * unit_price
        discount = 0
        total_formula_col = 4
        if ws.max_column >= 5 and (ws.cell(row, 5).value not in (None, "") or row_rate is not None):
            total_formula_col = 5
            if row_rate is not None:
                discount_value = ws.cell(row, 4).value
                computed_discount = formula_total(discount_value, quantity, unit_price, 0, row_rate)
                discount = computed_discount if computed_discount is not None else round(gross * row_rate)
        total_value = ws.cell(row, total_formula_col).value
        total = as_int(total_value) if not isinstance(total_value, str) else formula_total(total_value, quantity, unit_price, discount, row_rate)
        if total is None:
            total = gross - discount

        item = ParsedItem(first, quantity, unit_price, gross, discount, total, row_rate)
        items.append(item)
        if row_rate:
            line_discounts[discount_key_for_label(first)] = row_rate

    transport, notes, footer_discounts = parse_footer(ws, footer_start)
    return ParsedInvoice(
        path=path,
        client_name=client_name,
        order_date=order_date,
        secondary_line=secondary_line,
        transport=transport,
        notes=notes,
        footer_discounts=[] if line_discounts else footer_discounts,
        line_discounts_by_format=line_discounts,
        items=items,
    )


OFFERING_RE = re.compile(
    r"\s*(?P<label>(?:[1-9]\d*\s*x\s*)?\d+(?:[\.,]\d+)?\s*(?:kg|gr|g)|x\s*\d+(?:[\.,]\d+)?\s*(?:kg|gr|g)|[1-9]\d*\s*x\s*\d+)\s*$",
    flags=re.IGNORECASE,
)


PRODUCT_NAME_CORRECTIONS = {
    "aevna arrollada": "Avena Arrollada",
    "avea arrollada": "Avena Arrollada",
    "avena arrolada": "Avena Arrollada",
    "avena arrrollada": "Avena Arrollada",
    "avena instantane": "Avena Instantánea",
    "avena instantane a": "Avena Instantánea",
    "cebada pwrlada": "Cebada Perlada",
    "garnbanzos": "Garbanzos",
    "h. maiz blanca": "Harina de Maíz Blanca",
    "harina abati": "Harina de Maíz Abatí",
    "arroz 5/0": "Arroz 5/0 Largo Fino",
    "maiz parido blanco": "Maíz Partido Blanco",
    "maiz partdo blanco": "Maíz Partido Blanco",
    "maiz partido balnco": "Maíz Partido Blanco",
    "maiz patido blanco": "Maíz Partido Blanco",
    "maiz pisingalo": "Maíz Pisingallo",
    "mezcla para pajaros": "Mezcla para Pájaros",
    "porotos alubia": "Porotos Alubia",
    "semola de trigo": "Sémola de Trigo",
}


def normalize_product_name(name: str) -> str:
    cleaned = clean_cell_text(name)
    key = normalize_text(cleaned)
    if key in PRODUCT_NAME_CORRECTIONS:
        return PRODUCT_NAME_CORRECTIONS[key]
    if key == "arroz largo fino 5/0":
        return "Arroz 5/0 Largo Fino"
    if key == "arroz 5/0 largo fino":
        return "Arroz 5/0 Largo Fino"
    return cleaned


def normalize_offering_label(label: str) -> str:
    cleaned = clean_cell_text(label)
    normalized = normalize_text(cleaned).replace(",", ".")
    normalized = re.sub(r"\s*x\s*", "x", normalized)
    normalized = re.sub(r"\s+", " ", normalized)

    replacements = {
        "10x 1 kg": "10x1 kg",
        "10x1 kg": "10x1 kg",
        "10x5000 gr": "10x500 gr",
        "12x300 gr": "16x300 gr",
        "12x400 g": "12x400 gr",
        "12x400gr": "12x400 gr",
        "x25kg": "x 25 kg",
        "x30kg": "x 30 kg",
        "x1 kg": "10x1 kg",
        "x 1 kg": "10x1 kg",
        "x400 gr": "12x400 gr",
        "x4000 gr": "x 4 kg",
        "x500 gr": "10x500 gr",
        "x 500 gr": "10x500 gr",
        "x5 g": "x 5 kg",
        "x 5 g": "x 5 kg",
        "x5 kg": "x 5 kg",
        "x 5 kg": "x 5 kg",
    }
    if normalized in replacements:
        return replacements[normalized]

    pack_match = re.fullmatch(r"(\d+)x(\d+(?:\.\d+)?)\s*(kg|gr|g)", normalized)
    if pack_match:
        count, amount, unit = pack_match.groups()
        unit = "gr" if unit == "g" else unit
        amount = amount.rstrip("0").rstrip(".") if "." in amount else amount
        return f"{count}x{amount} {unit}"

    pack_without_unit = re.fullmatch(r"(\d+)x(\d+)", normalized)
    if pack_without_unit:
        count, amount = pack_without_unit.groups()
        return f"{count}x{amount} gr"

    bag_match = re.fullmatch(r"x(\d+(?:\.\d+)?)\s*(kg|gr|g)", normalized)
    if bag_match:
        amount, unit = bag_match.groups()
        unit = "gr" if unit == "g" else unit
        amount = amount.rstrip("0").rstrip(".") if "." in amount else amount
        label = f"x {amount} {unit}"
        return replacements.get(label, label)

    return cleaned


def split_raw_product_label(label: str) -> tuple[str, str]:
    cleaned = clean_cell_text(label)
    match = OFFERING_RE.search(cleaned)
    if not match:
        return cleaned, "Unidad"
    product_name = cleaned[: match.start()].strip()
    offering = match.group("label").strip()
    return product_name or cleaned, offering


def split_product_label(label: str) -> tuple[str, str]:
    product_name, offering = split_raw_product_label(label)
    return normalize_product_name(product_name), normalize_offering_label(offering)


def collect_source_files(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if not path.name.startswith("~$") and path.name.lower() != "totales.xlsx"
    )


def connect_database():
    from sqlalchemy import MetaData, create_engine

    url = os.getenv("GRANALIA_POSTGRES_URL", DEFAULT_POSTGRES_URL)
    engine = create_engine(url, future=True)
    metadata = MetaData()
    metadata.reflect(
        bind=engine,
        only=[
            "catalogs",
            "transports",
            "customers",
            "products",
            "product_offerings",
            "invoices",
            "invoice_items",
        ],
    )
    tables = metadata.tables
    return SimpleNamespace(
        engine=engine,
        catalogs=tables["catalogs"],
        transports=tables["transports"],
        customers=tables["customers"],
        products=tables["products"],
        product_offerings=tables["product_offerings"],
        invoices=tables["invoices"],
        invoice_items=tables["invoice_items"],
    )


def resolve_transport_id(repo, *, connection, transport_name: str | None, now) -> int | None:
    from sqlalchemy import select, update

    name = str(transport_name or "").strip()
    if not name:
        return None
    existing = connection.execute(
        select(repo.transports).where(repo.transports.c.name == name).order_by(repo.transports.c.transport_id).limit(1)
    ).mappings().first()
    if existing:
        transport_id = int(existing["transport_id"])
        connection.execute(update(repo.transports).where(repo.transports.c.transport_id == transport_id).values(updated_at=now))
        return transport_id
    return int(
        connection.execute(
            repo.transports.insert()
            .values(name=name, notes=[], created_at=now, updated_at=now)
            .returning(repo.transports.c.transport_id)
        ).scalar_one()
    )


def reset_database(repo) -> None:
    from sqlalchemy import text

    with repo.engine.begin() as connection:
        connection.execute(
            text(
                """
                TRUNCATE TABLE
                    invoice_items,
                    invoices,
                    catalogs,
                    product_offerings,
                    products,
                    customers,
                    transports
                RESTART IDENTITY CASCADE
                """
            )
        )


def import_invoices(source_dir: Path, dry_run: bool = False, reset_db: bool = False) -> dict[str, int]:
    parsed: list[ParsedInvoice] = []
    skipped = 0
    for path in collect_source_files(source_dir):
        try:
            invoice = parse_invoice(path)
        except Exception as exc:
            skipped += 1
            print(f"SKIP {path}: {exc}")
            continue
        if not invoice.items:
            skipped += 1
            print(f"SKIP {path}: sin items")
            continue
        parsed.append(invoice)
    parsed.sort(key=lambda invoice: (invoice.order_date, invoice.client_name, invoice.path.as_posix()))

    print(f"Facturas parseadas: {len(parsed)}; omitidas: {skipped}")
    if parsed:
        print(f"Rango de fechas: {parsed[0].order_date.isoformat()} a {parsed[-1].order_date.isoformat()}")
    if dry_run:
        customers = {invoice.client_name for invoice in parsed}
        transports = {invoice.transport for invoice in parsed if invoice.transport}
        products = {split_product_label(item.label)[0] for invoice in parsed for item in invoice.items}
        print(f"Clientes: {len(customers)}; transportes: {len(transports)}; productos: {len(products)}")
        return {"parsed": len(parsed), "skipped": skipped, "inserted": 0, "updated": 0}

    from sqlalchemy import select, update
    from sqlalchemy.dialects.postgresql import insert

    repo = connect_database()
    if reset_db:
        reset_database(repo)
        print("Base limpiada: clientes, transportes, productos, catalogos y facturas.")

    now = datetime.now(timezone.utc)
    inserted = 0
    updated = 0

    all_products: set[str] = set()
    latest_price: dict[tuple[str, str], tuple[date, int]] = {}
    aliases_by_product: dict[str, set[str]] = defaultdict(set)
    for invoice in parsed:
        for item in invoice.items:
            product_name, offering_label = split_product_label(item.label)
            raw_product_name, _raw_offering_label = split_raw_product_label(item.label)
            raw_product_name = clean_cell_text(raw_product_name)
            if raw_product_name and raw_product_name != product_name:
                aliases_by_product[product_name].add(raw_product_name)
            all_products.add(product_name)
            if offering_label not in ALLOWED_OFFERINGS:
                continue
            key = (product_name, offering_label)
            current = latest_price.get(key)
            if current is None or invoice.order_date >= current[0]:
                latest_price[key] = (invoice.order_date, item.unit_price)

    with repo.engine.begin() as connection:
        product_ids: dict[str, int] = {}
        offering_ids: dict[tuple[str, str], int] = {}

        for product_name in sorted(all_products):
            aliases = sorted(aliases_by_product.get(product_name, set()))
            stmt = insert(repo.products).values(
                name=product_name,
                aliases=aliases,
                active=True,
                created_at=now,
                updated_at=now,
            )
            existing = connection.execute(select(repo.products).where(repo.products.c.name == product_name)).mappings().first()
            if existing:
                product_id = int(existing["id"])
                connection.execute(update(repo.products).where(repo.products.c.id == product_id).values(aliases=aliases, active=True, updated_at=now))
            else:
                product_id = int(connection.execute(stmt.returning(repo.products.c.id)).scalar_one())
            product_ids[product_name] = product_id

        positions: dict[str, int] = {}
        for (product_name, offering_label), (_latest_date, price) in sorted(latest_price.items()):
            product_id = product_ids[product_name]
            positions[product_name] = positions.get(product_name, 0) + 1
            existing_offering = connection.execute(
                select(repo.product_offerings).where(
                    repo.product_offerings.c.product_id == product_id,
                    repo.product_offerings.c.label == offering_label,
                )
            ).mappings().first()
            if existing_offering:
                offering_id = int(existing_offering["id"])
                connection.execute(
                    update(repo.product_offerings)
                    .where(repo.product_offerings.c.id == offering_id)
                    .values(price=price, position=positions[product_name], active=True, updated_at=now)
                )
            else:
                offering_id = int(
                    connection.execute(
                        repo.product_offerings.insert()
                        .values(
                            product_id=product_id,
                            label=offering_label,
                            price=price,
                            position=positions[product_name],
                            active=True,
                            created_at=now,
                            updated_at=now,
                        )
                        .returning(repo.product_offerings.c.id)
                    ).scalar_one()
                )
            offering_ids[(product_name, offering_label)] = offering_id

        catalog = []
        for product_name, product_id in sorted(product_ids.items()):
            offerings = [
                {"id": offering_ids[(name, label)], "label": label, "price": price}
                for (name, label), (_latest_date, price) in sorted(latest_price.items())
                if name == product_name
            ]
            catalog.append({"id": product_id, "name": product_name, "aliases": sorted(aliases_by_product.get(product_name, set())), "offerings": offerings})
        connection.execute(update(repo.catalogs).where(repo.catalogs.c.active.is_(True)).values(active=False, updated_at=now))
        connection.execute(
            repo.catalogs.insert().values(
                name="Importacion 2026",
                active=True,
                source="import_2026",
                catalog=catalog,
                created_at=now,
                updated_at=now,
            )
        )

        customer_ids: dict[str, int] = {}
        transport_ids: dict[str, int | None] = {}
        profiles: dict[str, ParsedInvoice] = {}
        for invoice in parsed:
            profiles[invoice.client_name] = invoice

        for client_name, invoice in sorted(profiles.items()):
            transport_id = None
            if invoice.transport:
                transport_id = resolve_transport_id(repo, connection=connection, transport_name=invoice.transport, now=now)
                transport_ids[invoice.transport] = transport_id
            existing = connection.execute(select(repo.customers).where(repo.customers.c.name == client_name)).mappings().first()
            payload = {
                "name": client_name,
                "secondary_line": invoice.secondary_line,
                "notes": invoice.notes,
                "footer_discounts": invoice.footer_discounts,
                "line_discounts_by_format": invoice.line_discounts_by_format,
                "source_count": sum(1 for item in parsed if item.client_name == client_name),
                "transport_id": transport_id,
                "updated_at": now,
            }
            if existing:
                customer_id = int(existing["id"])
                connection.execute(update(repo.customers).where(repo.customers.c.id == customer_id).values(**payload))
            else:
                customer_id = int(connection.execute(repo.customers.insert().values(**payload, created_at=now).returning(repo.customers.c.id)).scalar_one())
            customer_ids[client_name] = customer_id

        for invoice_index, invoice in enumerate(parsed):
            invoice_created_at = datetime.combine(invoice.order_date, time.min, tzinfo=timezone.utc) + timedelta(seconds=invoice_index)
            transport_id = None
            if invoice.transport:
                transport_id = transport_ids.get(invoice.transport)
                if transport_id is None:
                    transport_id = resolve_transport_id(repo, connection=connection, transport_name=invoice.transport, now=invoice_created_at)
            item_payloads = []
            gross_total = 0
            final_total = 0
            total_bultos = 0
            for index, item in enumerate(invoice.items, start=1):
                product_name, offering_label = split_product_label(item.label)
                gross_total += item.gross
                final_total += item.total
                total_bultos += item.quantity
                item_payloads.append(
                    {
                        "line_number": index,
                        "product_id": product_ids.get(product_name),
                        "offering_id": offering_ids.get((product_name, offering_label)),
                        "label": item.label,
                        "quantity": item.quantity,
                        "unit_price": item.unit_price,
                        "gross": item.gross,
                        "discount": item.discount,
                        "total": item.total,
                    }
                )

            if invoice.footer_discounts:
                final_total = gross_total
                for discount in invoice.footer_discounts:
                    final_total -= round(final_total * float(discount["rate"]))

            existing_invoice_id = connection.execute(
                select(repo.invoices.c.id).where(repo.invoices.c.legacy_key == invoice.legacy_key)
            ).scalar_one_or_none()
            payload = {
                "customer_id": customer_ids.get(invoice.client_name),
                "transport_id": transport_id,
                "legacy_key": invoice.legacy_key,
                "client_name": invoice.client_name,
                "order_date": invoice.order_date,
                "secondary_line": invoice.secondary_line,
                "transport": invoice.transport,
                "notes": invoice.notes,
                "footer_discounts": invoice.footer_discounts,
                "line_discounts_by_format": invoice.line_discounts_by_format,
                "total_bultos": total_bultos,
                "gross_total": gross_total,
                "discount_total": gross_total - final_total,
                "final_total": final_total,
                "output_filename": invoice.path.name,
                "xlsx_data": invoice.path.read_bytes(),
                "xlsx_size": invoice.path.stat().st_size,
            }
            if existing_invoice_id:
                invoice_id = int(existing_invoice_id)
                connection.execute(update(repo.invoices).where(repo.invoices.c.id == invoice_id).values(**payload))
                connection.execute(repo.invoice_items.delete().where(repo.invoice_items.c.invoice_id == invoice_id))
                updated += 1
            else:
                invoice_id = int(connection.execute(repo.invoices.insert().values(**payload, created_at=invoice_created_at).returning(repo.invoices.c.id)).scalar_one())
                inserted += 1

            for item_payload in item_payloads:
                item_payload["invoice_id"] = invoice_id
            if item_payloads:
                connection.execute(repo.invoice_items.insert(), item_payloads)

    print(f"Importacion completa. Insertadas: {inserted}; actualizadas: {updated}; omitidas: {skipped}")
    return {"parsed": len(parsed), "skipped": skipped, "inserted": inserted, "updated": updated}


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa clientes, transportes, productos y facturas desde la carpeta 2026.")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR, help="Carpeta con archivos .xlsx de 2026")
    parser.add_argument("--dry-run", action="store_true", help="Parsea archivos y muestra conteos sin escribir en PostgreSQL")
    parser.add_argument(
        "--reset-db",
        action="store_true",
        help="Borra clientes, transportes, productos, catalogos y facturas antes de importar",
    )
    args = parser.parse_args()
    import_invoices(args.source_dir.resolve(), dry_run=args.dry_run, reset_db=args.reset_db)


if __name__ == "__main__":
    main()
