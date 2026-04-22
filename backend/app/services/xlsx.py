from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..domain.catalog import catalog_indexes
from ..domain.models import CatalogProduct, CustomerProfile, InvoiceRow, InvoiceSnapshot, InvoiceSummary, Order
from ..core.utils import clean_cell_text, derive_discount_mode, discount_key_for_label, safe_filename


LINE = Side(style="thin", color="767676")
SOFT_LINE = Side(style="thin", color="8D8D8D")
VERTICAL_LINE = Side(style="thin", color="808080")
NUMBER_FORMAT = "#,##0"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="ECECEC")
ALT_FILL = PatternFill(fill_type="solid", fgColor="F7F7F7")
WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")


def style_sheet(ws) -> None:
    ws.title = "adm"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["E"].hidden = True
    for row in range(1, 80):
        ws.row_dimensions[row].height = 22


def add_logo(ws, base_dir: Path) -> None:
    logo_path = base_dir / "img" / "logo.png"
    if not logo_path.exists():
        return
    image = XLImage(str(logo_path))
    image.width = 300
    image.height = 120
    ws.add_image(image, "A1")


def set_row_style(
    ws,
    row: int,
    size: int = 12,
    *,
    with_vertical_split: bool = False,
    fill: PatternFill | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> None:
    for col in range(1, 5):
        cell = ws.cell(row, col)
        cell.font = Font(size=size, color="333333")

        cell.border = Border(
            top=top if top else Side(style=None),
            bottom=bottom if bottom else Side(style=None),
            right=VERTICAL_LINE if with_vertical_split and col < 4 else Side(style=None),
        )

        if fill is not None:
            cell.fill = fill


def format_header(ws, base_dir: Path, order: Order, profile: CustomerProfile) -> int:
    add_logo(ws, base_dir)

    ws.merge_cells("C2:D3")
    ws["C2"] = "PRESUPUESTO"
    ws["C2"].font = Font(name="Cambria", size=18, bold=True, color="555555")
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

    row = 8
    set_row_style(ws, row, size=16, top=LINE, bottom=LINE)
    ws.merge_cells("A8:C8")
    ws["A8"] = f"Cliente:  {clean_cell_text(order.client_name)}"
    ws["A8"].font = Font(name="Cambria", size=16, bold=True, color="2F2F2F")
    ws["A8"].alignment = Alignment(horizontal="left")
    ws["D8"] = datetime.strptime(order.date, "%Y-%m-%d").strftime("%d/%m/%Y")
    ws["D8"].font = Font(name="Cambria", size=16, color="2F2F2F")
    ws["D8"].alignment = Alignment(horizontal="right")

    secondary_line = clean_cell_text(order.secondary_line or profile.secondary_line)
    if secondary_line:
        row += 1
        set_row_style(ws, row, size=14, bottom=LINE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1, secondary_line)
        ws.cell(row, 1).font = Font(name="Cambria", size=14, color="444444")
        ws.cell(row, 1).alignment = Alignment(horizontal="left")

    return row + 2


def header_table(ws, row: int) -> None:
    set_row_style(
        ws,
        row,
        size=14,
        with_vertical_split=True,
        fill=HEADER_FILL,
        top=LINE,
        bottom=LINE,
    )

    ws.cell(row, 1, "Producto")
    ws.cell(row, 2, "Cantidad")
    ws.cell(row, 3, "Precio Unitario")
    ws.cell(row, 4, "Total")

    for col in range(1, 5):
        cell = ws.cell(row, col)
        cell.font = Font(name="Cambria", size=14, bold=True, color="333333")
        cell.alignment = Alignment(horizontal="center" if col != 1 else "left", vertical="center")


def choose_rate(profile: CustomerProfile, discount_key: str) -> float:
    line_map = profile.line_discounts_by_format or {}
    if discount_key in line_map:
        return float(line_map[discount_key])
    return 0.0


def expand_rows(order: Order, profile: CustomerProfile, catalog: list[CatalogProduct]) -> list[InvoiceRow]:
    products_by_id, offerings_by_key, _aliases = catalog_indexes(catalog)
    rows: list[InvoiceRow] = []
    mode = derive_discount_mode(profile.to_data()["footer_discounts"], profile.line_discounts_by_format)
    for item in order.items:
        if not item.product_id or not item.offering_id:
            continue
        qty = int(item.quantity or 0)
        bonus_qty = int(item.bonus_quantity or 0)
        if qty <= 0 and bonus_qty <= 0:
            continue
        product_key = str(item.product_id)
        offering_key = (product_key, str(item.offering_id))
        product = products_by_id[product_key]
        offering = offerings_by_key[offering_key]
        label = f"{product['name']} {offering['label']}"
        rate = choose_rate(profile, discount_key_for_label(offering["label"]))

        def append_row(quantity: int, unit_price: int) -> None:
            gross = quantity * unit_price
            if mode in {"line_discount_net", "line_desc_factor"}:
                discount = round(gross * rate)
                total = gross - discount
            else:
                discount = 0
                total = gross
            rows.append(InvoiceRow(item.product_id, item.offering_id, label, quantity, unit_price, gross, discount, total))

        if qty > 0:
            append_row(qty, int(offering["price"]))
        if bonus_qty > 0:
            append_row(bonus_qty, 0)
    return rows


def build_invoice_snapshot(order: Order, profile: CustomerProfile, catalog: list[CatalogProduct]) -> InvoiceSnapshot:
    rows = expand_rows(order, profile, catalog)
    if not rows:
        raise ValueError("No hay productos cargados")
    summary = compute_summary(rows, profile)
    return InvoiceSnapshot(rows=rows, summary=summary, order=order, profile=profile)


def fill_product_rows(ws, rows: list[InvoiceRow], start_row: int) -> int:
    row = start_row
    for index, item in enumerate(rows):
        row_fill = ALT_FILL if index % 2 == 0 else WHITE_FILL
        set_row_style(
            ws,
            row,
            size=14,
            with_vertical_split=True,
            fill=row_fill,
            bottom=SOFT_LINE,
        )

        ws.cell(row, 1, item.label)
        ws.cell(row, 2, item.quantity)
        ws.cell(row, 3, item.unit_price)
        ws.cell(row, 4, item.total)

        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 3).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 4).alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(row, 2).number_format = NUMBER_FORMAT
        ws.cell(row, 3).number_format = NUMBER_FORMAT
        ws.cell(row, 4).number_format = NUMBER_FORMAT

        for col in range(1, 5):
            ws.cell(row, col).font = Font(name="Cambria", size=14, color="3A3A3A")

        row += 1
    return row - 1


def summary_discount_text(profile: CustomerProfile) -> str:
    discounts = profile.footer_discounts
    if not discounts:
        return ""
    parts = []
    for discount in discounts:
        parts.append(f"{round(float(discount.rate) * 100, 2):g}%")
    return " + ".join(parts)


def compute_summary(rows: list[InvoiceRow], profile: CustomerProfile) -> InvoiceSummary:
    gross_total = sum(int(item.gross) for item in rows)
    total_bultos = sum(int(item.quantity) for item in rows)
    mode = derive_discount_mode(profile.to_data()["footer_discounts"], profile.line_discounts_by_format)

    if mode in {"line_discount_net", "line_desc_factor"}:
        final_total = sum(int(item.total) for item in rows)
        discount_total = gross_total - final_total
    elif mode in {"summary_discount", "summary_multi_discount"}:
        running = gross_total
        for discount in profile.footer_discounts:
            running -= round(running * float(discount.rate))
        final_total = int(running)
        discount_total = gross_total - final_total
    else:
        final_total = gross_total
        discount_total = 0

    return InvoiceSummary(gross_total, discount_total, final_total, total_bultos)


def fill_footer(ws, start_row: int, profile: CustomerProfile, summary: InvoiceSummary, transport: str, notes: list[str]) -> None:
    row = start_row + 2

    for col in range(1, 5):
        ws.cell(row - 1, col).fill = WHITE_FILL
        ws.cell(row - 1, col).border = Border()

    # Línea separadora entre productos y resumen
    set_row_style(ws, row, size=14, with_vertical_split=True, fill=WHITE_FILL, top=LINE, bottom=SOFT_LINE)
    ws.cell(row, 1, "Total Bultos")
    ws.cell(row, 1).font = Font(name="Cambria", size=14, bold=True, color="333333")

    ws.cell(row, 2, summary.total_bultos)
    ws.cell(row, 2).font = Font(name="Cambria", size=14, bold=True, color="333333")
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    ws.cell(row, 2).number_format = NUMBER_FORMAT

    ws.cell(row, 4, summary.gross_total)
    ws.cell(row, 4).font = Font(name="Cambria", size=14, bold=True, color="333333")
    ws.cell(row, 4).alignment = Alignment(horizontal="right")
    ws.cell(row, 4).number_format = '$' + NUMBER_FORMAT
    row += 1

    if summary.discount_total > 0:
        set_row_style(ws, row, size=14, with_vertical_split=True, fill=WHITE_FILL, bottom=SOFT_LINE)
        discount_label = summary_discount_text(profile)
        ws.cell(row, 1, f"Dto {discount_label}" if discount_label else "Dto")
        ws.cell(row, 1).font = Font(name="Cambria", size=14, bold=True, color="333333")

        ws.cell(row, 4, summary.discount_total)
        ws.cell(row, 4).font = Font(name="Cambria", size=14, bold=True, color="333333")
        ws.cell(row, 4).alignment = Alignment(horizontal="right")
        ws.cell(row, 4).number_format = '$' + NUMBER_FORMAT
        row += 1

    set_row_style(ws, row, size=16, with_vertical_split=True, fill=WHITE_FILL, bottom=LINE)
    ws.cell(row, 1, "TOTAL")
    ws.cell(row, 1).font = Font(name="Cambria", size=18, bold=True, color="2F2F2F")

    ws.cell(row, 4, summary.final_total)
    ws.cell(row, 4).font = Font(name="Cambria", size=18, bold=True, color="2F2F2F")
    ws.cell(row, 4).alignment = Alignment(horizontal="right")
    ws.cell(row, 4).number_format = '$' + NUMBER_FORMAT
    row += 3

    if transport or notes:
        set_row_style(ws, row, size=14, bottom=LINE)
        row += 1

    if transport:
        transport_value = clean_cell_text(transport)
        transport_label = "Retira:"
        if transport_value.lower().startswith("retira"):
            transport_value = transport_value.split(":", 1)[-1].strip()

        ws.cell(row, 1, transport_label)
        ws.cell(row, 1).font = Font(name="Cambria", size=16, bold=True, color="333333")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2, transport_value)
        ws.cell(row, 2).font = Font(name="Cambria", size=16, color="333333")
        row += 1

    for note in notes:
        note_text = clean_cell_text(note)
        if not note_text:
            continue
        if note_text.lower().startswith("total de bultos") or note_text.lower().startswith("total bultos"):
            continue

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2, note_text)
        ws.cell(row, 2).font = Font(name="Cambria", size=14, color="333333")
        row += 1


def render_order_xlsx(base_dir: Path, order: Order, profile: CustomerProfile, catalog: list[CatalogProduct]) -> tuple[str, bytes]:
    workbook = Workbook()
    sheet = workbook.active
    style_sheet(sheet)
    header_row = format_header(sheet, base_dir, order, profile)
    header_table(sheet, header_row)

    snapshot = build_invoice_snapshot(order, profile, catalog)
    end_row = fill_product_rows(sheet, snapshot.rows, start_row=header_row + 1)
    fill_footer(
        sheet,
        end_row,
        profile,
        snapshot.summary,
        clean_cell_text(order.transport or profile.transport),
        [clean_cell_text(item) for item in (order.notes or profile.notes or []) if clean_cell_text(item)],
    )

    filename = f"{safe_filename(order.client_name)}{datetime.strptime(order.date, '%Y-%m-%d').strftime('%d-%m-%Y')}.xlsx"
    buffer = BytesIO()
    workbook.save(buffer)
    return filename, buffer.getvalue()


def export_order(base_dir: Path, order: Order, profile: CustomerProfile, catalog: list[CatalogProduct]) -> tuple[str, bytes]:
    return render_order_xlsx(base_dir, order, profile, catalog)
